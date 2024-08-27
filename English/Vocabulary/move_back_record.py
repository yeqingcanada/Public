from datetime import datetime
import pandas as pd
import re
from openpyxl import load_workbook

# cd 'C:\Users\Qing Ye\Ching-Notes\GitHub-Public-Code-And-Note\English\Vocabulary\'
# python move_back_record.py

file_path = r'C:\Users\Qing Ye\Ching-Notes\GitHub-Public-Code-And-Note\English\Vocabulary\ReadingListening.xlsx'
excel_file = pd.ExcelFile(file_path, engine='openpyxl')

# Regular expression pattern for date format YYYY-MM-DD
date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')

# Identify sheets with names matching the date format
date_sheets = [sheet for sheet in excel_file.sheet_names if date_pattern.match(sheet)]
date_sheets.sort()
combined_data = pd.DataFrame()

# -----------------------------------------------------------------------------------------------
# # Read data from the 'NEW' sheet if it exists
# if 'NEW' in excel_file.sheet_names:
#     combined_data = pd.read_excel(file_path, sheet_name='NEW', engine='openpyxl', header=None)
# else:
#     combined_data = pd.DataFrame()
# -----------------------------------------------------------------------------------------------

# Combine data from the identified sheets, ensuring column alignment
for sheet in date_sheets:
    df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl', header=None)
    combined_data = pd.concat([combined_data, df], ignore_index=True)

# Write the combined data back to the 'NEW' sheet
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    combined_data.to_excel(writer, sheet_name='NEW', index=False, header=False)

# Load the workbook to delete the date sheets
workbook = load_workbook(file_path)

# Delete the identified date sheets
for sheet in date_sheets:
    if sheet in workbook.sheetnames:
        workbook.remove(workbook[sheet])

# Save the workbook after removing sheets
workbook.save(file_path)

print("Data has been successfully combined into the 'NEW' sheet and the date sheets have been deleted.")
# ----------------------------------------------------------------------------------------------------------------

# Read the 'NEW' sheet and count the number of rows
new_df = pd.read_excel(file_path, sheet_name='NEW')
new_count = len(new_df)

# Read the 'record' sheet if it exists, otherwise create an empty DataFrame
if 'Record' in excel_file.sheet_names:
    record_df = pd.read_excel(file_path, sheet_name='Record')
else:
    record_df = pd.DataFrame(columns=['Date', 'Number of NEW', 'Interval Time', 'Record Changes'])

# Calculate the date difference and record change
if not record_df.empty:
    last_record_date = pd.to_datetime(record_df.iloc[-1]['Date'])
    days_since_last = (datetime.now() - last_record_date).days
    last_record_count = record_df.iloc[-1]['Number of NEW']
    record_change = new_count - last_record_count
else:
    days_since_last = None
    record_change = None

# Append the new record
new_record = {
    'Date': datetime.now().strftime('%Y-%m-%d'),
    'Number of NEW': new_count,
    'Interval Time': days_since_last,
    'Record Changes': record_change
}
record_df = record_df.append(new_record, ignore_index=True)

# Write the updated 'record' sheet back to the Excel file
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    record_df.to_excel(writer, sheet_name='Record', index=False)

print("Record updated successfully.")