import openpyxl
import sys
import io
import msvcrt
from datetime import datetime
from colorama import init, Fore, Style

# cd 'C:\Users\Qing Ye\Ching-Notes\GitHub-Public-Code-And-Note\English\Vocabulary\'
# python practice.py

path = r'C:\Users\Qing Ye\Ching-Notes\GitHub-Public-Code-And-Note\English\Vocabulary\ReadingListening.xlsx'
# path = r'C:\Users\Qing Ye\Ching-Notes\GitHub-Public-Code-And-Note\English\Vocabulary\SpeakingWriting.xlsx'

# 确保标准输出使用UTF-8编码
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')

# 读取Excel文件
wb = openpyxl.load_workbook(path)
new_sheet = wb['NEW']
already_sheet = wb['ALREADY']
no_need_sheet = wb['NO_NEED']

# 获取当前日期作为工作表名称
current_date_sheet_name = datetime.now().strftime('%Y-%m-%d')

# 检查是否存在名为当前日期的工作表，如果不存在则创建
if current_date_sheet_name in wb.sheetnames:
    not_yet_sheet = wb[current_date_sheet_name]
else:
    not_yet_sheet = wb.create_sheet(title=current_date_sheet_name)

# 获取NEW sheet的行数
new_rows = list(new_sheet.iter_rows(values_only=True))

# 将记录移动到目标sheet并删除原始行
def move_record(target_sheet, row_index):
    original_text, translated_text = new_rows[row_index]
    target_sheet.append([original_text, translated_text])
    new_sheet.delete_rows(row_index + 1)
    new_rows.pop(row_index)

# 互动逻辑
def interactive_session():
    i = 0
    while i < len(new_rows):
        original_text, translated_text = new_rows[i]

        init()

        # Your print statement
        print(f"Original Text {Fore.RED}({len(new_rows)}){Style.RESET_ALL}: {Fore.YELLOW}{original_text}{Style.RESET_ALL}", flush=True)
        # print(f"Original Text {Fore.RED}({len(new_rows)}){Style.RESET_ALL}:", flush=True)
        # print(f"{Fore.YELLOW}{original_text}{Style.RESET_ALL}", flush=True)

        print("Press space to show translation, 1 to keep in NEW, 2 to move to ALREADY, 3 to move to NO_NEED, or q to quit: ", end="", flush=True)
        action = msvcrt.getch().decode('utf-8')
        print(action)  # 打印用户输入的键以换行

        if action == ' ':
            # print(f"Translated Text: {Fore.BLUE}{translated_text}{Style.RESET_ALL}", flush=True)
            print(f"Translated Text: ", flush=True)
            print(f"{Fore.BLUE}{translated_text}{Style.RESET_ALL}", flush=True)
            print("Enter 1 to keep in NEW, 2 to move to ALREADY, 3 to move to NO_NEED, or q to quit: ", end="", flush=True)
            action = msvcrt.getch().decode('utf-8')
            print(action)  # 打印用户输入的键以换行

        if action == '1':
            move_record(not_yet_sheet, i)
        elif action == '2':
            move_record(already_sheet, i)
        elif action == '3':
            move_record(no_need_sheet, i)
        elif action == 'q':
            break
        else:
            print("Invalid input. Please try again.", flush=True)
            continue

    wb.save(path)
    print("Session complete. Changes saved.", flush=True)

# 开始互动会话
interactive_session()
